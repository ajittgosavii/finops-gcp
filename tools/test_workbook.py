"""The workbook's formulas must be right, not merely present.

openpyxl writes formula strings and never evaluates them, so "the file was
written" says nothing about whether =SUMIF points at the right rows. These tests
load the workbook into a real Excel calculation engine and compare every total
against the same Python model the deck is built from.

They also assert the two things that make the sheet honest: a blank rate produces
a dash rather than a zero (a zero reads as free), and the effort figures on the
Summary sheet agree with the slide.

    DATA_SOURCE=demo PYTHONPATH=services/api pytest tools
"""

from __future__ import annotations

import os
import shutil
import sys

import pytest

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, os.path.join(ROOT, "tools"))
sys.path.insert(0, os.path.join(ROOT, "packages", "finops-core", "src"))
sys.path.insert(0, os.path.join(ROOT, "services", "api"))
os.environ.setdefault("DATA_SOURCE", "demo")

pytest.importorskip("openpyxl")
formulas = pytest.importorskip("formulas")

import build_deck as bd  # noqa: E402
import build_workbook as bw  # noqa: E402

ONSHORE_RATE, OFFSHORE_RATE, DAYS = 1200, 350, 20


def _solve(path: str):
    """Evaluate the workbook. Keys look like '[name.xlsx]SHEET'!A1."""
    book = os.path.basename(path)
    sol = formulas.ExcelModel().loads(path).finish().calculate()

    def value(sheet: str, cell: str):
        return sol[f"'[{book}]{sheet.upper()}'!{cell}"].value[0, 0]

    return value


@pytest.fixture(scope="module")
def blank_rates(tmp_path_factory):
    out = str(tmp_path_factory.mktemp("wb") / "m.xlsx")
    bw.build(out)
    return _solve(out)


@pytest.fixture(scope="module")
def with_rates(tmp_path_factory):
    """A copy with day rates typed in, as a user would."""
    from openpyxl import load_workbook

    src = str(tmp_path_factory.mktemp("wb2") / "m.xlsx")
    bw.build(src)
    dst = str(tmp_path_factory.mktemp("wb3") / "rated.xlsx")
    shutil.copy(src, dst)
    wb = load_workbook(dst)
    wb["Inputs"]["B5"] = ONSHORE_RATE
    wb["Inputs"]["B6"] = OFFSHORE_RATE
    wb.save(dst)
    return _solve(dst)


def test_the_workbook_has_the_sheets_a_reader_expects() -> None:
    from openpyxl import load_workbook
    import tempfile

    out = os.path.join(tempfile.mkdtemp(), "m.xlsx")
    bw.build(out)
    assert load_workbook(out).sheetnames == ["Summary", "Inputs", "Effort", "Plan", "Gates"]


def test_the_effort_totals_evaluate_to_the_python_model(blank_rates) -> None:
    t = bd.effort_totals()
    v = blank_rates
    assert float(v("Summary", "B5")) == pytest.approx(t["onshore"])
    assert float(v("Summary", "B6")) == pytest.approx(t["offshore"])
    assert float(v("Summary", "B7")) == pytest.approx(t["base"])
    assert float(v("Summary", "B8")) == pytest.approx(t["offshore_pct"] / 100)
    assert float(v("Summary", "B9")) == pytest.approx(t["contingency"])
    assert float(v("Summary", "B10")) == pytest.approx(t["total"])
    assert float(v("Summary", "B11")) == pytest.approx(t["total"] * DAYS)


def test_peak_and_average_team_size_are_formulas_over_the_monthly_totals(blank_rates) -> None:
    months = [on + off for on, off in bd.effort_totals()["by_month"]]
    assert float(blank_rates("Summary", "B12")) == pytest.approx(max(months))
    assert float(blank_rates("Summary", "B13")) == pytest.approx(sum(months) / 4)


def test_a_blank_rate_shows_a_dash_not_a_zero(blank_rates) -> None:
    """A zero in a cost column reads as 'free'. It must read as 'not yet known'."""
    for cell in ("B14",):
        assert not isinstance(blank_rates("Summary", cell), (int, float))
    assert not isinstance(blank_rates("Effort", "I5"), (int, float))


def test_cost_is_person_days_times_the_right_rate(with_rates) -> None:
    t = bd.effort_totals()
    expected_base = t["onshore"] * DAYS * ONSHORE_RATE + t["offshore"] * DAYS * OFFSHORE_RATE
    assert float(with_rates("Effort", "I15")) == pytest.approx(expected_base)
    assert float(with_rates("Summary", "B14")) == pytest.approx(expected_base * (1 + bd.CONTINGENCY))


def test_the_plan_sheet_derives_duration_and_month(blank_rates) -> None:
    first = bd.PLAN[0][1][0]          # ("Kick-off...", 1, 2, "Onshore")
    assert float(blank_rates("Plan", "F5")) == pytest.approx(first[2] - first[1] + 1)
    assert blank_rates("Plan", "G5") == "Month 1"


def test_the_workbook_and_the_deck_describe_the_same_plan() -> None:
    """One source. If the deck grows an activity, the workbook grows a row."""
    from openpyxl import load_workbook
    import tempfile

    out = os.path.join(tempfile.mkdtemp(), "m.xlsx")
    bw.build(out)
    ws = load_workbook(out)["Plan"]

    # Stop at the first blank row. Below it sit the COUNTA/MAX summary formulas,
    # which also live in column B and would otherwise be counted as activities.
    rows = 0
    r = 5
    while ws.cell(row=r, column=2).value:
        rows += 1
        r += 1
    assert rows == sum(len(a) for _, a in bd.PLAN)

    # ...and the sheet's own COUNTA covers exactly those rows.
    assert ws.cell(row=r + 1, column=2).value == f"=COUNTA(B5:B{r - 1})"
