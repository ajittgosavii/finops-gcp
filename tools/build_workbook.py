"""Generate the effort-estimation workbook (.xlsx) with live formulas.

    DATA_SOURCE=demo PYTHONPATH=services/api python tools/build_workbook.py

The roles, the monthly FTE, the activities and the gates are read from
`build_deck.EFFORT`, `build_deck.PLAN` and `build_deck.GATES`, so the workbook,
the deck and the manual cannot disagree about the plan.

Everything computed is a real Excel formula, not a value baked in by Python. Open
it, change an FTE cell or a rate, and every total moves. A spreadsheet of frozen
numbers is a screenshot with extra steps.

**The rate cells ship EMPTY on purpose.** Day rates are commercial, and a rate we
invented would be the least defensible number in the pack. Enter your own on the
Inputs sheet and the cost sheets populate themselves. Until then the cost columns
show a dash, and they say so.
"""

from __future__ import annotations

import os
import sys

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, ROOT)
sys.path.insert(0, os.path.join(ROOT, "tools"))
sys.path.insert(0, os.path.join(ROOT, "packages", "finops-core", "src"))
sys.path.insert(0, os.path.join(ROOT, "services", "api"))
os.environ.setdefault("DATA_SOURCE", "demo")

from openpyxl import Workbook  # noqa: E402
from openpyxl.formatting.rule import CellIsRule  # noqa: E402
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side  # noqa: E402
from openpyxl.utils import get_column_letter  # noqa: E402
from openpyxl.worksheet.datavalidation import DataValidation  # noqa: E402

import build_deck as bd  # noqa: E402

INK = "FF0B142A"
BODY = "FF2E3A4E"
MUTED = "FF4E5766"
WASH = "FFF5F8FC"
RULE = "FFE2E7EF"
AZURE = "FF1E6FD9"
TEAL = "FF119B8A"
CRIMSON = "FFC23333"
AMBER = "FFC98500"
PAPER = "FFFFFFFF"

H1 = Font(name="Segoe UI", size=14, bold=True, color=INK)
H2 = Font(name="Segoe UI", size=10, bold=True, color=PAPER)
BOLD = Font(name="Segoe UI", size=10, bold=True, color=INK)
NORM = Font(name="Segoe UI", size=10, color=BODY)
SMALL = Font(name="Segoe UI", size=9, color=MUTED)
INPUT_FONT = Font(name="Segoe UI", size=10, bold=True, color="FF0000C0")

HEAD_FILL = PatternFill("solid", fgColor=INK)
WASH_FILL = PatternFill("solid", fgColor=WASH)
INPUT_FILL = PatternFill("solid", fgColor="FFFFF6D6")
THIN = Side(style="thin", color=RULE)
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

MONTHS = ["M1", "M2", "M3", "M4"]


def _title(ws, text: str, sub: str = "") -> None:
    ws["A1"] = text
    ws["A1"].font = H1
    if sub:
        ws["A2"] = sub
        ws["A2"].font = SMALL


def _header_row(ws, row: int, headers, widths=None) -> None:
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = H2
        c.fill = HEAD_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BOX
    if widths:
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w


# --------------------------------------------------------------------------
# 1. Inputs -- the only sheet a human types into
# --------------------------------------------------------------------------


def sheet_inputs(wb: Workbook):
    ws = wb.create_sheet("Inputs")
    _title(ws, "Inputs", "Yellow cells are yours. Everything else in this workbook is a formula.")

    rows = [
        ("Onshore day rate", None, "Your rate card. Leave blank and cost columns show a dash."),
        ("Offshore day rate", None, "Same."),
        ("Currency", "USD", "Label only; it does not convert anything."),
        ("Working days per month", 20, "Used to convert person-months to person-days."),
        ("Contingency", bd.CONTINGENCY, "Held by the delivery lead, not spent by default."),
        ("Duration (months)", 4, "Changing this does NOT re-plan the work. See the note below."),
    ]
    ws["A4"] = "Parameter"
    ws["B4"] = "Value"
    ws["C4"] = "Note"
    for c in ("A4", "B4", "C4"):
        ws[c].font = H2
        ws[c].fill = HEAD_FILL
        ws[c].border = BOX

    for i, (label, value, note) in enumerate(rows, start=5):
        ws.cell(row=i, column=1, value=label).font = BOLD
        c = ws.cell(row=i, column=2, value=value)
        c.font = INPUT_FONT
        c.fill = INPUT_FILL
        c.border = BOX
        ws.cell(row=i, column=3, value=note).font = SMALL
        if label == "Contingency":
            c.number_format = "0%"
        if "rate" in label:
            c.number_format = '#,##0;-#,##0;"—"'

    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 78

    ws["A13"] = "Read this before you change anything"
    ws["A13"].font = Font(name="Segoe UI", size=10, bold=True, color=CRIMSON)
    for i, t in enumerate([
        "Day rates ship EMPTY on purpose. A rate we invented would be the least defensible number in the pack.",
        "Effort is the estimate. Cost is arithmetic on top of it. Change the FTE cells on 'Effort', not the totals.",
        "Cutting 'Duration (months)' does not compress the plan. The long pole is access — credentials on four payers",
        "and four FOCUS exports — and no amount of staffing shortens it. See the 'Plan' sheet: G1 is a client gate.",
        "Contingency is held, not spent. It covers the OCI connector's first contact with a live tenancy, and the",
        "possibility that our KPI definitions disagree with Con Edison's and somebody has to find out who is right.",
    ], start=14):
        ws.cell(row=i, column=1, value=t).font = SMALL

    return ws


# --------------------------------------------------------------------------
# 2. Effort -- roles x months, with real formulas
# --------------------------------------------------------------------------


def sheet_effort(wb: Workbook):
    ws = wb.create_sheet("Effort")
    _title(ws, "Effort by role", "Person-months. Edit the FTE cells; every total below is a formula.")

    headers = ["Role", "Location"] + MONTHS + ["Person-months", "Person-days", "Cost", "Why staffed this way"]
    _header_row(ws, 4, headers, widths=[26, 11, 7, 7, 7, 7, 15, 13, 14, 62])

    dv = DataValidation(type="list", formula1='"Onshore,Offshore"', allow_blank=False)
    ws.add_data_validation(dv)

    first = 5
    for i, (role, loc, fte, why) in enumerate(bd.EFFORT):
        r = first + i
        ws.cell(row=r, column=1, value=role).font = NORM
        c = ws.cell(row=r, column=2, value=loc)
        c.font = Font(name="Segoe UI", size=10, bold=True, color=AZURE if loc == "Onshore" else TEAL)
        dv.add(c)
        for m in range(4):
            cell = ws.cell(row=r, column=3 + m, value=fte[m])
            cell.font = INPUT_FONT
            cell.fill = INPUT_FILL
            cell.number_format = "0.00"
            cell.alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=7, value=f"=SUM(C{r}:F{r})").number_format = "0.00"
        ws.cell(row=r, column=8, value=f"=G{r}*Inputs!$B$8").number_format = "0.0"
        # A rate that is blank must show a dash, not a zero. A zero reads as free.
        ws.cell(row=r, column=9,
                value=f'=IF(OR(Inputs!$B$5="",Inputs!$B$6=""),"—",'
                      f'H{r}*IF(B{r}="Onshore",Inputs!$B$5,Inputs!$B$6))').number_format = '#,##0;-#,##0;"—"'
        ws.cell(row=r, column=10, value=why).font = SMALL
        for col in range(1, 11):
            ws.cell(row=r, column=col).border = BOX

    last = first + len(bd.EFFORT) - 1
    tot = last + 1
    ws.cell(row=tot, column=1, value="TOTAL").font = BOLD
    for m in range(4):
        col = get_column_letter(3 + m)
        c = ws.cell(row=tot, column=3 + m, value=f"=SUM({col}{first}:{col}{last})")
        c.font = BOLD
        c.number_format = "0.00"
        c.alignment = Alignment(horizontal="center")
    for col, fmt in ((7, "0.0"), (8, "0"), (9, '#,##0;-#,##0;"—"')):
        L = get_column_letter(col)
        c = ws.cell(row=tot, column=col, value=f"=SUM({L}{first}:{L}{last})")
        c.font = BOLD
        c.number_format = fmt
    for col in range(1, 11):
        ws.cell(row=tot, column=col).fill = WASH_FILL
        ws.cell(row=tot, column=col).border = BOX
    ws.cell(row=tot, column=10, value="Peak team size is shown on the Summary sheet.").font = SMALL

    # A month whose team exceeds the peak we committed to should light up.
    ws.conditional_formatting.add(
        f"C{tot}:F{tot}",
        CellIsRule(operator="greaterThan", formula=["8.25"],
                   fill=PatternFill("solid", fgColor="FFFDE7E7"), font=Font(color=CRIMSON, bold=True)),
    )

    ws.freeze_panes = "A5"
    return ws, first, last, tot


# --------------------------------------------------------------------------
# 3. Summary
# --------------------------------------------------------------------------


def sheet_summary(wb: Workbook, first: int, last: int, tot: int):
    ws = wb.create_sheet("Summary", 0)
    _title(ws, "Multi-Cloud FinOps on Google Cloud — effort summary",
           "Every figure here is a formula over the Effort sheet. Nothing is typed.")

    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 76

    rows = [
        ("Onshore", f'=SUMIF(Effort!$B${first}:$B${last},"Onshore",Effort!$G${first}:$G${last})', "0.0",
         "Client-facing and judgement work: architecture, reconciliation, security."),
        ("Offshore", f'=SUMIF(Effort!$B${first}:$B${last},"Offshore",Effort!$G${first}:$G${last})', "0.0",
         "Build and test."),
        ("Base effort", "=B5+B6", "0.0", "Person-months."),
        ("Offshore share", "=IF(B7=0,0,B6/B7)", "0%", "A 70/30 split is the intent."),
        ("Contingency", "=B7*Inputs!$B$9", "0.0", "Held by the delivery lead. Not spent by default."),
        ("Total effort", "=B7+B9", "0.0", "Person-months, including contingency."),
        ("Total person-days", "=B10*Inputs!$B$8", "0", "At the working-days-per-month on the Inputs sheet."),
        ("Peak team size (FTE)", f"=MAX(Effort!C{tot}:F{tot})", "0.00", "Month 3. No month exceeds it."),
        ("Average team size (FTE)", f"=AVERAGE(Effort!C{tot}:F{tot})", "0.00", "Across the four months."),
        ("Total cost", f'=IF(OR(Inputs!$B$5="",Inputs!$B$6=""),"—",Effort!I{tot}*(1+Inputs!$B$9))',
         '#,##0;-#,##0;"—"', "Enter both day rates on the Inputs sheet. Includes contingency."),
    ]
    ws["A4"] = "Measure"
    ws["B4"] = "Value"
    ws["C4"] = "Note"
    for c in ("A4", "B4", "C4"):
        ws[c].font = H2
        ws[c].fill = HEAD_FILL
        ws[c].border = BOX

    for i, (label, formula, fmt, note) in enumerate(rows, start=5):
        ws.cell(row=i, column=1, value=label).font = BOLD
        c = ws.cell(row=i, column=2, value=formula)
        c.number_format = fmt
        c.font = NORM
        c.border = BOX
        c.alignment = Alignment(horizontal="right")
        ws.cell(row=i, column=3, value=note).font = SMALL

    ws["A16"] = "Team size by month"
    ws["A16"].font = BOLD
    for m in range(4):
        ws.cell(row=17, column=2 + m, value=MONTHS[m]).font = BOLD
        col = get_column_letter(3 + m)
        c = ws.cell(row=18, column=2 + m, value=f"=Effort!{col}{tot}")
        c.number_format = "0.00"
        c.font = NORM
        c.alignment = Alignment(horizontal="center")
    ws["A18"] = "FTE"
    ws["A18"].font = NORM

    ws["A20"] = "What this workbook does not tell you"
    ws["A20"].font = Font(name="Segoe UI", size=10, bold=True, color=CRIMSON)
    for i, t in enumerate([
        "Day rates. They ship empty on purpose; enter your own and every cost cell populates.",
        "That the plan can be compressed by adding people. It cannot: the first gate is client-side access.",
        "That the OCI connector has run against a live tenancy. It has not. Some of the contingency is for that.",
        "That our KPI definitions match Con Edison's. Month 3 exists partly to find out where they do not.",
    ], start=21):
        ws.cell(row=i, column=1, value=t).font = SMALL

    return ws


# --------------------------------------------------------------------------
# 4. Plan -- activities, weeks, gates
# --------------------------------------------------------------------------


def sheet_plan(wb: Workbook):
    ws = wb.create_sheet("Plan")
    _title(ws, "Delivery plan", "Weeks are inclusive. Duration and month are formulas over start and end.")

    headers = ["Workstream", "Activity", "Owner", "Start week", "End week", "Weeks", "Starts in", "Gate"]
    _header_row(ws, 4, headers, widths=[22, 46, 11, 11, 11, 9, 11, 34])

    gate_by_week = {w: (tag, what) for w, tag, what in bd.GATES}

    r = 5
    for group, acts in bd.PLAN:
        for name, w0, w1, owner in acts:
            ws.cell(row=r, column=1, value=group).font = NORM
            ws.cell(row=r, column=2, value=name).font = NORM
            c = ws.cell(row=r, column=3, value=owner)
            c.font = Font(name="Segoe UI", size=10, bold=True, color=AZURE if owner == "Onshore" else TEAL)
            for col, v in ((4, w0), (5, w1)):
                cc = ws.cell(row=r, column=col, value=v)
                cc.font = INPUT_FONT
                cc.fill = INPUT_FILL
                cc.alignment = Alignment(horizontal="center")
            ws.cell(row=r, column=6, value=f"=E{r}-D{r}+1").alignment = Alignment(horizontal="center")
            ws.cell(row=r, column=7, value=f'="Month "&ROUNDUP(D{r}/4,0)').alignment = Alignment(horizontal="center")
            gate = gate_by_week.get(w1)
            ws.cell(row=r, column=8, value=f"{gate[0]} — {gate[1]}" if gate else "").font = (
                Font(name="Segoe UI", size=9, bold=True, color=CRIMSON) if gate else SMALL
            )
            for col in range(1, 9):
                ws.cell(row=r, column=col).border = BOX
            r += 1

    # An end week before its start week is a data-entry error, and it should shout.
    ws.conditional_formatting.add(
        f"F5:F{r - 1}",
        CellIsRule(operator="lessThan", formula=["1"],
                   fill=PatternFill("solid", fgColor="FFFDE7E7"), font=Font(color=CRIMSON, bold=True)),
    )

    ws.cell(row=r + 1, column=1, value="Activities").font = BOLD
    ws.cell(row=r + 1, column=2, value=f"=COUNTA(B5:B{r - 1})").font = BOLD
    ws.cell(row=r + 2, column=1, value="Longest activity (weeks)").font = BOLD
    ws.cell(row=r + 2, column=2, value=f"=MAX(F5:F{r - 1})").font = BOLD
    ws.cell(row=r + 3, column=1, value="Plan length (weeks)").font = BOLD
    ws.cell(row=r + 3, column=2, value=f"=MAX(E5:E{r - 1})").font = BOLD

    ws.freeze_panes = "A5"
    return ws


def sheet_gates(wb: Workbook):
    ws = wb.create_sheet("Gates")
    _title(ws, "Gates", "Nothing downstream starts before the gate above it clears.")
    _header_row(ws, 4, ["Gate", "Week", "Month", "What must be true", "Who owns it"], widths=[9, 9, 11, 62, 16])
    owners = {
        "G1": "Con Edison",
        "G2": "Joint",
        "G3": "Infosys",
        "G4": "Joint",
    }
    for i, (week, tag, what) in enumerate(bd.GATES, start=5):
        ws.cell(row=i, column=1, value=tag).font = Font(name="Segoe UI", size=10, bold=True, color=CRIMSON)
        ws.cell(row=i, column=2, value=week).alignment = Alignment(horizontal="center")
        ws.cell(row=i, column=3, value=f'="Month "&ROUNDUP(B{i}/4,0)').alignment = Alignment(horizontal="center")
        ws.cell(row=i, column=4, value=what).font = NORM
        ws.cell(row=i, column=5, value=owners.get(tag, "")).font = BOLD
        for col in range(1, 6):
            ws.cell(row=i, column=col).border = BOX
    ws["A11"] = ("G1 is a Con Edison gate, and it is the one that moves the end date. Read access on four payers "
                 "and four FOCUS exports. No amount of staffing shortens it.")
    ws["A11"].font = SMALL
    return ws


def build(out: str) -> str:
    wb = Workbook()
    wb.remove(wb.active)

    sheet_inputs(wb)
    _, first, last, tot = sheet_effort(wb)
    sheet_summary(wb, first, last, tot)   # inserted at index 0
    sheet_plan(wb)
    sheet_gates(wb)

    wb.active = 0
    wb.save(out)
    return out


def main() -> None:
    out = os.path.join(ROOT, "docs", "Infosys_FinOps_GCP_Effort_Model.xlsx")
    build(out)
    print("wrote", os.path.relpath(out, ROOT))


if __name__ == "__main__":
    main()
